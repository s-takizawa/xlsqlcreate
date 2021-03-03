#!python3
import openpyxl
import sys
import os.path as osp
from datetime import datetime

# constants
TEMPLATE_FILE_DIR = 'template_data'  # directory name include template file
TEMPLATE_FILE_NAME = 'template_sqlcreate.xlsx'  # sql create template file name
SHEET_NAME = 'template'  # sheet name at sql create template file
RESULT_FILE_DIR = 'result_data'  # directory name include result file
RESULT_FILE_NAME = 'result_{0}.txt'  # output result txt name

BASE_SQL_ROW = 3
BASE_SQL_COL = 2

DATA_START_ROW = 3
DATA_START_COL = 5

# method


def end_proc():
    """process end
    """
    input("please,press any key to end")
    exit()


def main():
    """main process
    """
    # operation
    print('start create sql statement.')

    current_dir = osp.dirname(sys.argv[0])
    template_xlsx_file_path = osp.join(
        current_dir, TEMPLATE_FILE_DIR, TEMPLATE_FILE_NAME)
    # check file path
    if not osp.exists(template_xlsx_file_path):
        print('ERROR! can not found template xlsx file[{0:s}].'.format(
            template_xlsx_file_path))
        return

    # open template elsx file
    wb = openpyxl.load_workbook(template_xlsx_file_path, data_only=True)
    ws = wb[SHEET_NAME]

    # get base sql statement
    base_sql = ws.cell(row=BASE_SQL_ROW, column=BASE_SQL_COL).value
    if len(base_sql) == 0:
        print('ERROR! not found base sql statement[{0:s}].'.format(
            template_xlsx_file_path))
        return

    # end cell of data range
    data_end_row = ws.max_row
    data_end_col = ws.max_column
    if DATA_START_ROW > data_end_row or DATA_START_COL > data_end_col:
        print('ERROR! Invalid data[{0:s}].'.format(template_xlsx_file_path))
        return

    # output file path
    res_fname = RESULT_FILE_NAME.format(
        datetime.now().strftime('%Y%m%d_%H%M%S'))
    res_file_path = osp.join(current_dir, RESULT_FILE_DIR, res_fname)

    # create/write
    with open(res_file_path, 'w', encoding='utf-8') as f:
        # loop row unit
        for r in range(DATA_START_ROW, data_end_row + 1):
            input_data_list = []
            # loop column unit
            for c in range(DATA_START_COL, data_end_col + 1):
                temp_val = str(ws.cell(row=r, column=c).value) if not ws.cell(
                    row=r, column=c).value is None else str("")
                input_data_list.append(temp_val)

            if len(input_data_list) > 0:
                # check data: Skip if all characters are blank
                if len(list(filter(lambda val: len(val) > 0, input_data_list))) == 0:
                    continue

                # write sql statement
                sql_stmt = base_sql.format(*input_data_list)
                f.write(sql_stmt + '\n')

    wb.close()
    print('success!!!')


if __name__ == "__main__":
    main()
    end_proc()
